import openpyxl
wb = openpyxl.load_workbook('total.xlsx')
ws = wb.active
ws = wb['data']
name_list = []

for row in ws.iter_rows(min_row=2, min_col=2) :
    print(row[0].value)
    if row[0].value not in name_list :
        name_list.append(row[0].value)
        wb.create_sheet(row[0].value)
    # 데이터는 row에 있고 list data에 담을 예정
    data = []   # list를 담을 data 공간
    for cell in row : 
        data.append(cell.value)     # data를 추가하기 위한 구문
    wb[row[0].value].append(data)   # 선택한 row에 data를 추가

# print(name_list)

# for name in name_list :
#     wb.create_sheet(name)
wb.save('total.xlsx')