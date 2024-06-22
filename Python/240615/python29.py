import openpyxl

total_wb = openpyxl.Workbook()
total_ws = total_wb.active
total_ws.title = "data"

total_ws.append(["순번", "제품명", "가격", "수량", "합계"])

file_list = ['A업체.xlsx', 'B업체.xlsx', 'C업체.xlsx']


# 통합작업 시작
for file in file_list :
    wb = openpyxl.load_workbook(file, data_only=True)
    ws = wb.active
    for row in ws.iter_rows(min_row = 2) :
        data = []
        for cell in row :
            data.append(cell.value)
        total_ws.append(data)

# 통합파일 순번 변경
# count = 1
# for row in total_ws.iter_rows(min_row=2, max_col=1) :
#     for cell in row :
#         print(cell.value)
#         cell.value = count
#         count = count + 1
#         print(count)

# for row in total_ws.iter_rows(min_row=2, max_col=1) :
#     for cell in row :
#         cell.value = row[0].row-1

i=0
for cell in total_ws['A'] :
    if i!=0 :
        cell.value=i
    i+=1

# 엑셀 통합작업
total_wb.save('total.xlsx')


