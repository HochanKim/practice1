import openpyxl
# save_path = 'result/A업체.xlsx'     # result/하위폴더
save_path = 'B업체.xlsx'
wb = openpyxl.load_workbook(save_path, data_only=True)      # data_only -> 연산된 결과로 읽어오게 하는 불값 기능

ws = wb['data']

for x in range(1, 8+1) :
    for y in range(1, 5+1) :
        # print(str(x)+":"+str(y), end=" ")
        print(ws.cell(row=x, column=y).value, end=" ")
    print()

# for row in ws.iter_rows() :
#     print(row)

# for row in ws.iter_rows(min_row=2, max_row=5) :
#     print(row)

for row in ws.iter_rows(min_row=2, max_row=7) :
    for cell in row :
        print(cell.value, end = " ")
    print()