import openpyxl
wb = openpyxl.Workbook()

sheet = wb.active
sheet = wb['Sheet']     # Sheet라는 이름의 시트를 선택한다

# sheet['A1'] = "hello"
# sheet['B1'] = "world"

sheet['A1'] = '날짜'
sheet['B1'] = '제품명'
sheet['C1'] = '가격'
sheet['D1'] = '수량'
sheet['E1'] = '합계'

# 두 번째 행
sheet.cell(row = 2, column = 1, value = '2023-01-02')
sheet.cell(row = 2, column = 2, value = '마우스')
sheet.cell(row = 2, column = 3, value = 5000)
sheet.cell(row = 2, column = 4, value = 302)
sheet.cell(row = 2, column = 5, value = '=C2*D2')

# 세 번째 행
sheet.append(['2030-01-09', '키보드', 7000, 35, '=C3*D3'])

# 수정
sheet['C2'] = 4000
sheet['D2'] = 40

# 삭제
del sheet['A3']

# 열
rows = [
    ['2030-01-09', '키보드', 7000, 35, '=C4*D4'],
    ['2023-01-16', '키보드', 7000, 35, '=C5*D5'],
    ['2025-04-09', '키보드', 7000, 35, '=C6*D6'],
    ['2032-07-10', '키보드', 7000, 35, '=C7*D7']
]

count = 4
for row in rows :
    print(row[4])
    print(f'=C{count}*D{count}')
    row[4]=f'=C{count}*D{count}'
    sheet.append(row)
    count=count+1

# 모든 행 읽어오기
for row in sheet.iter_rows(values_only=True) :
    if(row[1]=='마우스') :
        print(row)

wb.save('세번째파일.xlsx')