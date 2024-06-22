import openpyxl
wb = openpyxl.load_workbook('total.xlsx')
ws = wb.active
ws = wb['data']
name_list = []

for row in ws.iter_rows(min_row=2, min_col=2) :
    print(row[0].value)
    if row[0].value not in name_list :
        name_list.append(row[0].value)
print()
print(name_list)

for name in name_list :
    wb.create_sheet(name)
wb.save('total.xlsx')